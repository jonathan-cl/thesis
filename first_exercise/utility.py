from openpyxl import load_workbook
from scipy.interpolate import interp1d
import numpy as np
from plotly.subplots import make_subplots
import plotly.graph_objects as go


def get_survival_prob(s, T):
    wb = load_workbook("first_exercise\\lifetable.xlsx")
    ws = wb["Sheet1"]
    # male death probabilities from 2019, see https://www.ssa.gov/oact/STATS/table4c6.html
    p_survival = [1-c[0].value for c in ws["B3:B122"]]
    wb.close()
    return p_survival[s:T]


def get_annuity_payment(annuity_investment, start_age, end_age, R):
    p_survival = get_survival_prob(start_age, end_age)
    accumulated_p_surv = 1
    res = 0
    for t in range(end_age-start_age):
        accumulated_p_surv *= p_survival[t]
        res += accumulated_p_surv / R**t
    return annuity_investment / res


# For now only allow annuity investments that are a multiple of 5 (are on the grid)
def simulate(c, w0, grid_w, an_investment, an, r):
    an_i = an_investment // 5  # Position in grid
    n_periods = c.shape[0]

    # Init solution vectors
    sim_c = np.nan * np.zeros(n_periods)
    sim_w = np.nan * np.zeros(n_periods)

    sim_w[0] = w0 - an_investment
    sim_c[0] = interp1d(grid_w, c[0][an_i])(sim_w[0])
    for t in range(1, n_periods):
        sim_w[t] = (sim_w[t-1] - sim_c[t-1] + an) * r
        sim_c[t] = interp1d(grid_w, c[t][an_i], fill_value="extrapolate")(sim_w[t])
    return sim_c, sim_w


def plot_simulation(sim_c, sim_w, s, T, w_0, an_inv):
    fig = make_subplots()
    fig.add_trace(go.Scatter(x=np.arange(s, T), y=sim_c, name="consumption"))
    fig.add_trace(go.Scatter(x=np.arange(s, T), y=sim_w, name="wealth"))
    annotations = [dict(xref='paper', yref='paper', x=0.0, y=1.05,
                            xanchor='left', yanchor='bottom',
                            text='w_0 = {}, annuity investment = {}'.format(w_0, an_inv),
                            font=dict(family='Arial',
                                      size=30,
                                      color='rgb(37,37,37)'),
                            showarrow=False)]
    fig.update_xaxes(title_text="Age")
    fig.update_yaxes(title_text="Consumption / Wealth")
    fig.update_layout(annotations=annotations)
    fig.show()

