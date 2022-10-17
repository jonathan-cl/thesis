from openpyxl import load_workbook
from scipy.interpolate import interp1d
import numpy as np
from plotly import make_subplots
import plotly.graph_objects as go


def get_survival_prob(s, T):
    wb = load_workbook("lifetable.xlsx")
    ws = wb["Sheet1"]
    # male death probabilities from 2019, see https://www.ssa.gov/oact/STATS/table4c6.html
    p_survival = [1-c[0].value for c in ws["B3:B122"]]
    wb.close()
    return p_survival[s:T]


def get_annuity_payment(start_age, end_age, r):
    p_survival = get_survival_prob(start_age, end_age)
    accumulated_p_surv = 1
    res = 0
    for t in range(end_age-start_age):
        accumulated_p_surv *= p_survival[t]
        res += accumulated_p_surv / r**t
    return 1 / res


# For now only allow annuity investments that are a multiple of 5 (are on the grid)
def simulate(c, w0, grid_w, an_investment, an, r):
    n_periods = c.shape[0]
    sim_c = np.nan * np.zeros(n_periods)
    sim_m = np.nan * np.zeros(n_periods)
    sim_m[0] = w0 - an_investment
    sim_c[0] = interp1d(grid_w, c[0][int(an_investment/5)])(sim_m[0])
    for t in range(1, n_periods):
        sim_m[t] = (sim_m[t-1] - sim_c[t-1] + an) * r
        if sim_m[t] < 0.001:
            print(t)
        sim_c[t] = interp1d(grid_w, c[t][int(an_investment / 5)], fill_value="extrapolate")(sim_m[t])
    return sim_c, sim_m


def plot_simulation(sim_c, sim_m, s, T, w_0, an_inv):
    fig = make_subplots()
    fig.add_trace(go.Scatter(x=np.arange(s, T), y=sim_c, name="consumption"))
    fig.add_trace(go.Scatter(x=np.arange(s, T), y=sim_m, name="wealth"))
    annotations = [dict(xref='paper', yref='paper', x=0.0, y=1.05,
                            xanchor='left', yanchor='bottom',
                            text='w_0 = {}, annuity investment = {}'.format(w_0, an_inv),
                            font=dict(family='Arial',
                                      size=30,
                                      color='rgb(37,37,37)'),
                            showarrow=False)]
    fig.update_xaxes(title_text="Age")
    fig.update_yaxes(title_text="Consumption / Wealth")
    fig.update_layout(annotations=annotations)
    fig.show()

